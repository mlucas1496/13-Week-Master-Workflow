"""
Balance Updater Web Application

A local web app to process USDx balance files and update Activity Rollforward files.
"""

from flask import Flask, render_template, request, send_file, jsonify
import os
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula
from datetime import datetime
from werkzeug.utils import secure_filename
import traceback
import re
import stacked_activity_updater
import fva_data_updater

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 1024 * 1024 * 1024  # 1GB max total upload size (for all 7 files combined)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def adjust_formula_columns(formula, column_offset=1):
    """
    Adjust column references in an Excel formula by a given offset.

    Example:
        Input: "=CC14+7", offset=1
        Output: "=CD14+7"

        Input: "=INDEX($B:$B,MATCH(CC14,$A:$A,0))", offset=1
        Output: "=INDEX($B:$B,MATCH(CD14,$A:$A,0))"

    Args:
        formula: Excel formula string starting with '='
        column_offset: Number of columns to shift right (default 1)

    Returns:
        Adjusted formula string
    """
    if not formula or not formula.startswith('='):
        return formula

    def col_to_num(col_letters):
        """Convert column letters to number (A=1, Z=26, AA=27, etc.)"""
        col_num = 0
        for char in col_letters:
            col_num = col_num * 26 + (ord(char) - ord('A') + 1)
        return col_num

    def num_to_col(num):
        """Convert number to column letters (1=A, 26=Z, 27=AA, etc.)"""
        col_letters = ''
        while num > 0:
            num -= 1
            col_letters = chr(ord('A') + (num % 26)) + col_letters
            num //= 26
        return col_letters

    def adjust_cell_ref(match):
        """Adjust a single cell reference"""
        prefix_char = match.group(1) or ''
        dollar1 = match.group(2) or ''
        col_letters = match.group(3)
        dollar2 = match.group(4) or ''
        row_number = match.group(5) or ''

        # Only adjust if column reference is relative (no $ before column)
        if dollar1:
            # Absolute column reference - don't adjust
            return prefix_char + dollar1 + col_letters + dollar2 + row_number
        else:
            # Relative column reference - adjust it
            col_num = col_to_num(col_letters)
            new_col = num_to_col(col_num + column_offset)
            return prefix_char + dollar1 + new_col + dollar2 + row_number

    # Pattern matches cell references with row numbers (not bare column letters)
    # This avoids matching function names like INDEX, MATCH, etc.
    # Matches: A1, $A1, A$1, $A$1, CC14, $CC$14, etc.
    # Also matches ranges like CC14:CC20
    pattern = r'(^|[^A-Za-z])(\$?)([A-Z]+)(\$?)(\d+)'

    # First pass: adjust cell references with row numbers
    adjusted_formula = re.sub(pattern, adjust_cell_ref, formula)

    # Second pass: adjust column-only ranges like $B:$B or A:Z
    def adjust_col_range(match):
        prefix = match.group(1) or ''
        dollar1 = match.group(2) or ''
        col1 = match.group(3)
        colon = match.group(4)
        dollar2 = match.group(5) or ''
        col2 = match.group(6)
        suffix = match.group(7) or ''

        # Adjust first column if relative
        if dollar1:
            new_col1 = col1
        else:
            new_col1 = num_to_col(col_to_num(col1) + column_offset)

        # Adjust second column if relative
        if dollar2:
            new_col2 = col2
        else:
            new_col2 = num_to_col(col_to_num(col2) + column_offset)

        return prefix + dollar1 + new_col1 + colon + dollar2 + new_col2 + suffix

    # Match column ranges like $B:$B, A:Z, etc.
    range_pattern = r'([^A-Za-z])(\$?)([A-Z]+)(:)(\$?)([A-Z]+)([^A-Za-z]|$)'
    adjusted_formula = re.sub(range_pattern, adjust_col_range, adjusted_formula)

    return adjusted_formula

def process_files(weekly_file_path, rollforward_file_path, bth_file_path=None, aggregator_file_path=None, fva_files=None):
    """
    Main processing function:
    1. Read column K (Current Week) from USDx Balances tab (K10 header)
    2. Find last date in row 14 of Beginning Balances tab
    3. Paste column K data into the next column
    4. Copy formula from row 10 to generate new date header
    5. Process Cockpit tab (copy formulas from last column to next)
    6. If BTH file provided, match date and paste Total BTH Financing to Cockpit row 20
    7. If Activity Aggregator provided, append filtered data to Stacked Activity tab
    8. If FvA files provided, update 1-Week/4-Week/13-Week FvA Data Tabs
    """

    log = []

    try:
        log.append("="*60)
        log.append("STEP 1: Reading Column K from USDx Balances")
        log.append("="*60)

        # Open source file
        wb_source = openpyxl.load_workbook(weekly_file_path, data_only=True)
        sheet_source = wb_source["USDx Balances"]

        # Column K is column 11, header in row 10
        balance_col = 11  # Column K
        ref_col = 2  # Column B for matching
        header_row_source = 10  # K10 contains "Current Week"
        data_start_row = 13

        balance_header = sheet_source.cell(header_row_source, balance_col).value
        log.append(f"Column K header (K10): {balance_header}")

        # Extract Current Week balances (column K) with reference numbers
        balances = {}
        for row in range(data_start_row, sheet_source.max_row + 1):
            ref_num = sheet_source.cell(row, ref_col).value
            balance = sheet_source.cell(row, balance_col).value

            if ref_num is not None:
                ref_str = str(ref_num).strip()
                # Skip section headers
                if not any(keyword in ref_str.lower() for keyword in ['bullish', 'coindesk', 'fiat', 'balance', 'total']):
                    try:
                        balance_val = float(balance) if balance is not None else 0.0
                    except (ValueError, TypeError):
                        balance_val = 0.0
                    balances[ref_str] = balance_val

        log.append(f"Extracted {len(balances)} Current Week balances from column K")

        wb_source.close()

        log.append("")
        log.append("="*60)
        log.append("STEP 2: Opening Activity Rollforward file")
        log.append("="*60)

        # Open target file (preserve formulas)
        log.append("Loading workbook...")
        wb_target = openpyxl.load_workbook(rollforward_file_path, keep_vba=False)

        # Also load once with data_only=True and read_only for faster reading of calculated values
        log.append("Loading calculated values (read-only mode)...")
        wb_target_data = openpyxl.load_workbook(rollforward_file_path, data_only=True, read_only=True, keep_vba=False)

        # Find Beginning Balances sheet
        if "Beginning Balances" in wb_target.sheetnames:
            sheet_target = wb_target["Beginning Balances"]
            log.append("Found 'Beginning Balances' sheet")
        else:
            return {
                'success': False,
                'log': log,
                'error': "Beginning Balances sheet not found in Activity Rollforward file"
            }

        log.append("")
        log.append("="*60)
        log.append("STEP 3: Finding last date in row 14")
        log.append("="*60)

        # Find the last date in row 14
        last_col_with_date = 0
        for col in range(1, sheet_target.max_column + 1):
            cell_val = sheet_target.cell(14, col).value
            if cell_val is not None:
                last_col_with_date = col

        if last_col_with_date == 0:
            return {
                'success': False,
                'log': log,
                'error': "No dates found in row 14 of Beginning Balances"
            }

        log.append(f"Last date found in column {openpyxl.utils.get_column_letter(last_col_with_date)} (row 14)")

        # The new column is one to the right
        new_col = last_col_with_date + 1
        log.append(f"Will paste into column {openpyxl.utils.get_column_letter(new_col)}")

        log.append("")
        log.append("="*60)
        log.append("STEP 4: Copying formatting from previous column")
        log.append("="*60)

        # Copy formatting from previous column to new column
        from copy import copy
        source_col = last_col_with_date
        log.append(f"Copying formatting from column {openpyxl.utils.get_column_letter(source_col)} to {openpyxl.utils.get_column_letter(new_col)}...")

        for row in range(1, sheet_target.max_row + 1):
            source_cell = sheet_target.cell(row, source_col)
            target_cell = sheet_target.cell(row, new_col)

            # Copy all formatting attributes
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

        # Copy column width
        source_col_letter = openpyxl.utils.get_column_letter(last_col_with_date)
        target_col_letter = openpyxl.utils.get_column_letter(new_col)
        sheet_target.column_dimensions[target_col_letter].width = sheet_target.column_dimensions[source_col_letter].width

        log.append(f"✓ Copied formatting and column width from {source_col_letter}")

        log.append("")
        log.append("="*60)
        log.append("STEP 5: Copying formula from row 10 for date header")
        log.append("="*60)

        # Copy formula from row 10 (previous column) to generate new date
        formula_source_cell = sheet_target.cell(10, last_col_with_date)
        formula_target_cell = sheet_target.cell(10, new_col)

        # Check if source has a formula (skip array formulas)
        if (formula_source_cell.value and
            not isinstance(formula_source_cell.value, ArrayFormula) and
            isinstance(formula_source_cell.value, str) and
            formula_source_cell.value.startswith('=')):
            # Adjust column references in formula
            try:
                adjusted_formula = adjust_formula_columns(formula_source_cell.value, column_offset=1)
                if isinstance(adjusted_formula, str) and adjusted_formula.startswith('='):
                    formula_target_cell.value = adjusted_formula
                    log.append(f"Copied formula from {openpyxl.utils.get_column_letter(last_col_with_date)}10 to {openpyxl.utils.get_column_letter(new_col)}10")
                else:
                    log.append(f"Formula adjustment failed for row 10")
            except Exception:
                log.append(f"Error adjusting formula in row 10")
        else:
            # If no formula, just note it
            log.append(f"No valid formula found in row 10, column {openpyxl.utils.get_column_letter(last_col_with_date)}")

        log.append("")
        log.append("="*60)
        log.append("STEP 6: Pasting Current Week balances")
        log.append("="*60)

        # Match and paste balances by Column B
        ref_col_target = 2  # Column B
        data_start_row_target = 15  # Data starts at row 15
        matched = 0
        not_matched = 0
        copied_from_prev = 0

        # Special references that should copy from previous week instead of Weekly Balances file
        copy_from_previous_refs = ['11', '60', '76', '91']

        for row in range(data_start_row_target, sheet_target.max_row + 1):
            ref_num = sheet_target.cell(row, ref_col_target).value

            if ref_num is not None:
                ref_str = str(ref_num).strip()

                # Check if this is a special reference that should copy from previous week
                # Only applies to rows between 15 and 148
                if ref_str in copy_from_previous_refs and row <= 148:
                    # Copy value from previous week's column
                    prev_value = sheet_target.cell(row, last_col_with_date).value
                    sheet_target.cell(row, new_col).value = prev_value
                    copied_from_prev += 1
                    log.append(f"  Row {row} (Ref {ref_str}): Copied {prev_value} from previous week")
                elif ref_str in balances:
                    sheet_target.cell(row, new_col).value = balances[ref_str]
                    matched += 1
                else:
                    sheet_target.cell(row, new_col).value = 0.0
                    not_matched += 1

        log.append(f"✓ Matched and pasted: {matched} records")
        log.append(f"✓ Copied from previous week: {copied_from_prev} records (refs: 11, 60, 76, 91)")
        if not_matched > 0:
            log.append(f"⚠ Not found in source: {not_matched} records (set to 0.0)")

        log.append("")
        log.append("="*60)
        log.append("STEP 7: Copying and adjusting formulas from previous week")
        log.append("="*60)

        # Copy formulas/cells from previous column (optimized)
        prev_col = last_col_with_date
        formulas_copied = 0
        formulas_adjusted = 0

        # Copy rows 12-20
        for row in range(12, 21):  # 12 to 20 inclusive
            source_cell = sheet_target.cell(row, prev_col)

            # Skip empty cells and array formulas
            if source_cell.value is None or isinstance(source_cell.value, ArrayFormula):
                continue

            target_cell = sheet_target.cell(row, new_col)

            # Copy formula or value
            if isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                # Adjust column references in formula
                try:
                    adjusted_formula = adjust_formula_columns(source_cell.value, column_offset=1)
                    if isinstance(adjusted_formula, str) and adjusted_formula.startswith('='):
                        target_cell.value = adjusted_formula
                        formulas_copied += 1
                        if adjusted_formula != source_cell.value:
                            formulas_adjusted += 1
                    else:
                        continue
                except Exception:
                    continue
            else:
                target_cell.value = source_cell.value

        # Copy rows 148-239
        formulas_copied_148 = 0
        for row in range(148, 240):  # 148 to 239 inclusive
            source_cell = sheet_target.cell(row, prev_col)

            # Skip empty cells and array formulas
            if source_cell.value is None or isinstance(source_cell.value, ArrayFormula):
                continue

            target_cell = sheet_target.cell(row, new_col)

            # Copy formula or value
            if isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                # Adjust column references in formula
                try:
                    adjusted_formula = adjust_formula_columns(source_cell.value, column_offset=1)
                    if isinstance(adjusted_formula, str) and adjusted_formula.startswith('='):
                        target_cell.value = adjusted_formula
                        formulas_copied_148 += 1
                        if adjusted_formula != source_cell.value:
                            formulas_adjusted += 1
                    else:
                        continue
                except Exception:
                    continue
            else:
                target_cell.value = source_cell.value

        log.append(f"✓ Copied {formulas_copied + formulas_copied_148} formulas, {formulas_adjusted} adjusted")

        log.append("")
        log.append("="*60)
        log.append("STEP 8: Processing Cockpit Tab")
        log.append("="*60)

        # Check if Cockpit tab exists
        if "Cockpit" in wb_target.sheetnames:
            sheet_cockpit = wb_target["Cockpit"]
            log.append("✓ Found Cockpit tab")

            # Use pre-loaded workbook with calculated values to find most recent date
            sheet_cockpit_temp = wb_target_data["Cockpit"]

            # Find the column with the MOST RECENT DATE in ROW 10
            most_recent_date = None
            most_recent_date_col = None

            log.append(f"Scanning Cockpit row 10 for most recent date...")

            for col in range(1, sheet_cockpit_temp.max_column + 1):
                cell_val = sheet_cockpit_temp.cell(10, col).value

                # Check if this cell contains a date
                if cell_val and isinstance(cell_val, datetime):
                    if most_recent_date is None or cell_val > most_recent_date:
                        most_recent_date = cell_val
                        most_recent_date_col = col

            if most_recent_date_col:
                log.append(f"✓ Most recent date: {most_recent_date.strftime('%Y-%m-%d')} in column {openpyxl.utils.get_column_letter(most_recent_date_col)}")
            else:
                log.append(f"⚠ No dates found in Cockpit row 10")

            if most_recent_date_col:
                cockpit_prev_col = most_recent_date_col
                cockpit_new_col = most_recent_date_col + 1

                log.append("")
                log.append(f"Will copy ENTIRE column {openpyxl.utils.get_column_letter(cockpit_prev_col)} → {openpyxl.utils.get_column_letter(cockpit_new_col)}")
                log.append(f"Copying column {openpyxl.utils.get_column_letter(cockpit_prev_col)} → {openpyxl.utils.get_column_letter(cockpit_new_col)}")

                # Copy formatting from previous column
                from copy import copy
                log.append("Copying formatting from source column...")
                for row in range(1, sheet_cockpit.max_row + 1):
                    source_cell = sheet_cockpit.cell(row, cockpit_prev_col)
                    target_cell = sheet_cockpit.cell(row, cockpit_new_col)

                    # Copy formatting
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.number_format = copy(source_cell.number_format)
                        target_cell.protection = copy(source_cell.protection)
                        target_cell.alignment = copy(source_cell.alignment)

                # Copy column width
                cockpit_source_letter = openpyxl.utils.get_column_letter(cockpit_prev_col)
                cockpit_target_letter = openpyxl.utils.get_column_letter(cockpit_new_col)
                sheet_cockpit.column_dimensions[cockpit_target_letter].width = sheet_cockpit.column_dimensions[cockpit_source_letter].width

                log.append("✓ Formatting copied")

                # Copy formulas/values from previous column to new column (optimized)
                log.append("Copying formulas and values...")
                cockpit_formulas_copied = 0
                cockpit_values_copied = 0

                # Only iterate up to actual used rows (not max_row which can be inflated)
                max_row = min(sheet_cockpit.max_row, 1000)  # Cap at 1000 rows for performance

                for row in range(1, max_row + 1):
                    source_cell = sheet_cockpit.cell(row, cockpit_prev_col)

                    # Skip empty cells and array formulas
                    if source_cell.value is None or isinstance(source_cell.value, ArrayFormula):
                        continue

                    target_cell = sheet_cockpit.cell(row, cockpit_new_col)

                    # Copy formula or value
                    if isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                        # Adjust column references in formula
                        try:
                            adjusted_formula = adjust_formula_columns(source_cell.value, column_offset=1)
                            if isinstance(adjusted_formula, str) and adjusted_formula.startswith('='):
                                target_cell.value = adjusted_formula
                                cockpit_formulas_copied += 1
                            else:
                                continue
                        except Exception:
                            continue
                    else:
                        target_cell.value = source_cell.value
                        cockpit_values_copied += 1

                log.append(f"✓ Copied {cockpit_formulas_copied} formulas to Cockpit column {openpyxl.utils.get_column_letter(cockpit_new_col)}")
                log.append(f"✓ Copied {cockpit_values_copied} values to Cockpit column {openpyxl.utils.get_column_letter(cockpit_new_col)}")

                # Process BTH file if provided
                if bth_file_path:
                    log.append("")
                    log.append("="*60)
                    log.append("STEP 9: Processing BTH Investments File")
                    log.append("="*60)

                    try:
                        wb_bth = openpyxl.load_workbook(bth_file_path, data_only=True)

                        if "Summary" in wb_bth.sheetnames:
                            sheet_bth = wb_bth["Summary"]
                            log.append("✓ Found Summary tab in BTH file")

                            # Find the MOST RECENT DATE in row 30 of BTH Summary tab
                            bth_most_recent_date = None
                            bth_most_recent_col = None

                            log.append("Scanning BTH row 30 for dates...")
                            for col in range(1, sheet_bth.max_column + 1):
                                bth_date = sheet_bth.cell(30, col).value

                                if bth_date and isinstance(bth_date, datetime):
                                    if bth_most_recent_date is None or bth_date > bth_most_recent_date:
                                        bth_most_recent_date = bth_date
                                        bth_most_recent_col = col

                            if bth_most_recent_col:
                                # Get Total BTH Financing from row 42 for most recent date
                                bth_financing_total = sheet_bth.cell(42, bth_most_recent_col).value
                                log.append(f"✓ Most recent BTH date: {bth_most_recent_date.strftime('%Y-%m-%d')}")
                                log.append(f"✓ Found in column {openpyxl.utils.get_column_letter(bth_most_recent_col)}")
                                log.append(f"✓ Total BTH Financing (row 42): {bth_financing_total}")

                                # Check what's currently in Cockpit row 20 before pasting
                                current_val = sheet_cockpit.cell(20, cockpit_new_col).value
                                log.append(f"Current value in Cockpit {openpyxl.utils.get_column_letter(cockpit_new_col)}20: {current_val}")

                                # Paste to Cockpit tab, row 20, new column
                                sheet_cockpit.cell(20, cockpit_new_col).value = bth_financing_total
                                log.append(f"✓ Pasted {bth_financing_total} to Cockpit {openpyxl.utils.get_column_letter(cockpit_new_col)}20")

                                # VERIFY the value was pasted correctly
                                verification_value = sheet_cockpit.cell(20, cockpit_new_col).value
                                if verification_value == bth_financing_total:
                                    log.append(f"✅ VERIFIED: Cell {openpyxl.utils.get_column_letter(cockpit_new_col)}20 = {verification_value}")
                                else:
                                    log.append(f"⚠️ WARNING: Verification failed! Expected {bth_financing_total}, got {verification_value}")
                            else:
                                log.append("⚠ Could not find any dates in BTH file row 30")

                        else:
                            log.append("⚠ Summary tab not found in BTH file")

                        wb_bth.close()

                    except Exception as e:
                        log.append(f"⚠ Error processing BTH file: {str(e)}")
                        log.append(traceback.format_exc())
                else:
                    log.append("")
                    log.append("STEP 9: BTH file not provided - skipping BTH integration")
            else:
                log.append("⚠ Could not find any date columns in Cockpit tab")
        else:
            log.append("⚠ Cockpit tab not found in Activity Rollforward file")

        log.append("")
        log.append("="*60)
        log.append("STEP 11: Processing Rollforward Check Tab")
        log.append("="*60)

        # Check if Rollforward Check tab exists
        if "Rollforward Check" in wb_target.sheetnames:
            sheet_rollforward_check = wb_target["Rollforward Check"]
            log.append("✓ Found Rollforward Check tab")

            # STEP 11A: Find column with "cumulative" in row 10 (BEFORE any insertions)
            cumulative_col_original = None
            log.append("Searching for 'cumulative' in row 10...")

            for col in range(1, sheet_rollforward_check.max_column + 1):
                cell_val = sheet_rollforward_check.cell(10, col).value
                if cell_val and isinstance(cell_val, str) and "cumulative" in cell_val.lower():
                    cumulative_col_original = col
                    col_letter = openpyxl.utils.get_column_letter(col)
                    log.append(f"✓ Found 'cumulative' in column {col_letter}10: {cell_val}")
                    break

            if cumulative_col_original:
                # STEP 11B: Find the column with greatest value in row 8 (BEFORE any insertions)
                # Use pre-loaded workbook with calculated values
                sheet_rollforward_temp = wb_target_data["Rollforward Check"]

                max_value = None
                max_value_col_original = None

                log.append("Scanning row 8 for greatest numeric value...")
                for col in range(1, sheet_rollforward_temp.max_column + 1):
                    cell_val = sheet_rollforward_temp.cell(8, col).value

                    # Only consider numeric values (integers or floats)
                    if isinstance(cell_val, (int, float)) and cell_val is not None:
                        if max_value is None or cell_val > max_value:
                            max_value = cell_val
                            max_value_col_original = col

                if max_value_col_original:
                    log.append(f"✓ Greatest value in row 8: {max_value} in column {openpyxl.utils.get_column_letter(max_value_col_original)}")

                    # STEP 11C: Insert 6 columns before "Cumulative"
                    log.append(f"Inserting 6 columns before {openpyxl.utils.get_column_letter(cumulative_col_original)} (Cumulative)...")
                    sheet_rollforward_check.insert_cols(cumulative_col_original, 6)
                    log.append(f"✓ Inserted 6 columns. 'Cumulative' moved to column {openpyxl.utils.get_column_letter(cumulative_col_original + 6)}")

                    # STEP 11D: Determine if source columns were affected by insertion
                    # If max_value_col was before cumulative, it's unaffected
                    # If it was at or after cumulative, it shifted right by 6
                    if max_value_col_original >= cumulative_col_original:
                        source_start_col = max_value_col_original + 6
                        log.append(f"Source columns shifted right by 6: {openpyxl.utils.get_column_letter(max_value_col_original)} → {openpyxl.utils.get_column_letter(source_start_col)}")
                    else:
                        source_start_col = max_value_col_original
                        log.append(f"Source columns unaffected by insertion: {openpyxl.utils.get_column_letter(source_start_col)}")

                    # STEP 11E: Copy 6 columns from source to newly inserted columns
                    # Source: columns starting at max value (6 consecutive columns)
                    # Target: the 6 newly inserted columns (cumulative_col_original to cumulative_col_original+5)
                    source_end_col = source_start_col + 5  # 6 columns total (0-5)
                    target_start_col = cumulative_col_original
                    target_end_col = cumulative_col_original + 5

                    # Calculate offset for formula adjustment
                    column_offset = target_start_col - source_start_col

                    log.append(f"Copying 6 columns: {openpyxl.utils.get_column_letter(source_start_col)}-{openpyxl.utils.get_column_letter(source_end_col)} → {openpyxl.utils.get_column_letter(target_start_col)}-{openpyxl.utils.get_column_letter(target_end_col)}")
                    log.append(f"Formula adjustment offset: {column_offset} columns")

                    # Copy formatting and values
                    from copy import copy
                    cells_copied = 0
                    formulas_copied = 0
                    array_formulas_skipped = 0

                    # Process reasonable number of rows
                    max_row = min(sheet_rollforward_check.max_row, 1000)

                    log.append("Copying formatting and values...")

                    for row in range(1, max_row + 1):
                        for i in range(6):  # 6 columns
                            source_col = source_start_col + i
                            target_col = target_start_col + i

                            source_cell = sheet_rollforward_check.cell(row, source_col)
                            target_cell = sheet_rollforward_check.cell(row, target_col)

                            # Copy formatting
                            if source_cell.has_style:
                                target_cell.font = copy(source_cell.font)
                                target_cell.border = copy(source_cell.border)
                                target_cell.fill = copy(source_cell.fill)
                                target_cell.number_format = copy(source_cell.number_format)
                                target_cell.protection = copy(source_cell.protection)
                                target_cell.alignment = copy(source_cell.alignment)

                            # Skip empty cells for value copy
                            if source_cell.value is None:
                                continue

                            # Skip array formulas (openpyxl can't write these properly)
                            if isinstance(source_cell.value, ArrayFormula):
                                array_formulas_skipped += 1
                                continue

                            # Copy formula with adjustment or copy value
                            if isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                                # Validate and adjust formula
                                try:
                                    adjusted_formula = adjust_formula_columns(source_cell.value, column_offset=column_offset)
                                    # Only write if it's a valid string formula
                                    if isinstance(adjusted_formula, str) and adjusted_formula.startswith('='):
                                        target_cell.value = adjusted_formula
                                        formulas_copied += 1
                                    else:
                                        # Invalid formula, skip
                                        continue
                                except Exception:
                                    # Formula adjustment failed, skip this cell
                                    continue
                            else:
                                # Copy non-formula value
                                target_cell.value = source_cell.value

                            cells_copied += 1

                    # Copy column widths
                    for i in range(6):
                        source_col_letter = openpyxl.utils.get_column_letter(source_start_col + i)
                        target_col_letter = openpyxl.utils.get_column_letter(target_start_col + i)
                        if source_col_letter in sheet_rollforward_check.column_dimensions:
                            sheet_rollforward_check.column_dimensions[target_col_letter].width = sheet_rollforward_check.column_dimensions[source_col_letter].width

                    if array_formulas_skipped > 0:
                        log.append(f"⚠ Skipped {array_formulas_skipped} array formulas (not supported)")

                    log.append(f"✓ Copied {cells_copied} cells, {formulas_copied} formulas adjusted")

                    # STEP 11F: Copy formulas from MQ to MW (including array formulas)
                    mq_col = openpyxl.utils.column_index_from_string('MQ')
                    mw_col = openpyxl.utils.column_index_from_string('MW')

                    # Check if columns exist
                    if mq_col <= sheet_rollforward_check.max_column and mw_col <= sheet_rollforward_check.max_column:
                        log.append(f"Copying formulas and formatting from MQ → MW (rows 14-97)...")

                        from copy import copy
                        mw_formulas_copied = 0

                        # First copy formatting for entire column
                        for row in range(1, max_row + 1):
                            source_cell = sheet_rollforward_check.cell(row, mq_col)
                            target_cell = sheet_rollforward_check.cell(row, mw_col)

                            # Copy formatting
                            if source_cell.has_style:
                                target_cell.font = copy(source_cell.font)
                                target_cell.border = copy(source_cell.border)
                                target_cell.fill = copy(source_cell.fill)
                                target_cell.number_format = copy(source_cell.number_format)
                                target_cell.protection = copy(source_cell.protection)
                                target_cell.alignment = copy(source_cell.alignment)

                        # Then copy formulas for rows 14-97
                        for row in range(14, 98):  # Rows 14-97
                            source_cell = sheet_rollforward_check.cell(row, mq_col)

                            if source_cell.value is None:
                                continue

                            target_cell = sheet_rollforward_check.cell(row, mw_col)

                            # Handle ArrayFormula - extract text and convert to regular formula
                            if isinstance(source_cell.value, ArrayFormula):
                                if hasattr(source_cell.value, 'text') and source_cell.value.text:
                                    formula_text = source_cell.value.text
                                    # Replace MQ$ references with MW$
                                    adjusted_formula = formula_text.replace('MQ$', 'MW$')
                                    # Ensure it starts with =
                                    if not adjusted_formula.startswith('='):
                                        adjusted_formula = '=' + adjusted_formula
                                    target_cell.value = adjusted_formula
                                    mw_formulas_copied += 1
                            # Handle regular formulas
                            elif isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                                # Replace MQ$ references with MW$
                                adjusted_formula = source_cell.value.replace('MQ$', 'MW$')
                                target_cell.value = adjusted_formula
                                mw_formulas_copied += 1
                            # Copy other values
                            else:
                                target_cell.value = source_cell.value

                        # Copy column width
                        if 'MQ' in sheet_rollforward_check.column_dimensions:
                            sheet_rollforward_check.column_dimensions['MW'].width = sheet_rollforward_check.column_dimensions['MQ'].width

                        log.append(f"✓ Copied {mw_formulas_copied} formulas and formatting from MQ → MW")
                    else:
                        log.append(f"⚠ Columns MQ or MW not found")

                    # STEP 11F: Set column widths for MT through MX (16.67)
                    # After insertion, MT is now at position cumulative_col_original + 6
                    # MT through MX is 5 columns total
                    mt_position = cumulative_col_original + 6
                    mx_position = mt_position + 4  # MX is 4 columns after MT

                    log.append(f"Setting column widths MT-MX to 16.67...")
                    for col_offset in range(5):  # MT, MU, MV, MW, MX (5 columns)
                        col = mt_position + col_offset
                        col_letter = openpyxl.utils.get_column_letter(col)
                        sheet_rollforward_check.column_dimensions[col_letter].width = 16.67

                    log.append(f"✓ Set width 16.67 for columns {openpyxl.utils.get_column_letter(mt_position)}-{openpyxl.utils.get_column_letter(mx_position)}")

                else:
                    log.append("⚠ Could not find any numeric values in Rollforward Check row 8")
            else:
                log.append("⚠ Could not find 'cumulative' in Rollforward Check row 10")
        else:
            log.append("⚠ Rollforward Check tab not found in Activity Rollforward file")

        # Final verification before saving
        log.append("")
        log.append("="*60)
        log.append("PRE-SAVE VERIFICATION")
        log.append("="*60)

        # Check if Cockpit was processed
        cockpit_processed = "Cockpit" in wb_target.sheetnames and 'most_recent_date_col' in locals() and most_recent_date_col is not None

        if cockpit_processed:
            sheet_cockpit = wb_target["Cockpit"]
            cockpit_new_col = most_recent_date_col + 1

            log.append(f"Checking Cockpit column {openpyxl.utils.get_column_letter(cockpit_new_col)} before save:")

            # Sample a few important rows
            row_10_val = sheet_cockpit.cell(10, cockpit_new_col).value
            row_20_val = sheet_cockpit.cell(20, cockpit_new_col).value

            log.append(f"  Row 10: {repr(row_10_val)[:60]}")
            log.append(f"  Row 20: {repr(row_20_val)[:60]}")

            # Check if any data exists in the new column
            has_data = False
            data_count = 0
            for row in range(1, min(sheet_cockpit.max_row + 1, 100)):
                if sheet_cockpit.cell(row, cockpit_new_col).value is not None:
                    has_data = True
                    data_count += 1

            if has_data:
                log.append(f"✅ Cockpit column {openpyxl.utils.get_column_letter(cockpit_new_col)} has {data_count} cells with data")
            else:
                log.append(f"⚠️ WARNING: Cockpit column {openpyxl.utils.get_column_letter(cockpit_new_col)} appears to be empty!")
        else:
            log.append("Cockpit tab was not processed (not found or no date column identified)")

        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Activity_Rollforward_Updated_{timestamp}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)

        log.append("")
        log.append("="*60)
        log.append("STEP 10: Saving output file")
        log.append("="*60)

        try:
            wb_target.save(output_path)
            log.append(f"✓ Saved to: {output_filename}")
        except Exception as save_error:
            log.append(f"✗ ERROR saving file: {str(save_error)}")
            log.append(traceback.format_exc())
            raise

        wb_target.close()
        wb_target_data.close()

        log.append("")
        log.append("="*60)
        log.append("SUMMARY")
        log.append("="*60)
        log.append(f"✓ Beginning Balances: Column pasted to {openpyxl.utils.get_column_letter(new_col)}")
        log.append(f"✓ Beginning Balances: {matched} balances matched")
        log.append(f"✓ Beginning Balances: {formulas_copied + formulas_copied_148} formulas copied")

        if "Cockpit" in wb_target.sheetnames and 'most_recent_date_col' in locals() and most_recent_date_col:
            log.append(f"✓ Cockpit: Copied column {openpyxl.utils.get_column_letter(most_recent_date_col)} → {openpyxl.utils.get_column_letter(most_recent_date_col + 1)}")

        if "Rollforward Check" in wb_target.sheetnames and 'cumulative_col_original' in locals() and cumulative_col_original:
            log.append(f"✓ Rollforward Check: Inserted 6 columns before Cumulative")
            if 'cells_copied' in locals():
                log.append(f"✓ Rollforward Check: Copied {cells_copied} cells, {formulas_copied} formulas")
            if 'mw_formulas_copied' in locals():
                log.append(f"✓ Rollforward Check: Copied {mw_formulas_copied} formulas from MQ → MW")
            log.append(f"✓ Rollforward Check: Set column widths MT-MX to 16.67")

        # Process Stacked Activity tab if Activity Aggregator file provided
        if aggregator_file_path:
            log.append("="*60)
            log.append("STEP 11: Updating Stacked Activity Tab")
            log.append("="*60)

            sa_result = stacked_activity_updater.append_activity_data(
                aggregator_file_path,
                output_path,
                output_path,  # Update in place
                log
            )

            if not sa_result['success']:
                log.append(f"⚠ Warning: Stacked Activity update failed: {sa_result.get('error', 'Unknown error')}")
                log.append("  Continuing with other processing...")

        # Process FvA Data Tabs if 13WCF files provided
        if fva_files:
            log.append("="*60)
            log.append("STEP 12: Updating FvA Data Tabs")
            log.append("="*60)

            fva_result = fva_data_updater.update_fva_tabs(
                output_path,
                fva_files,
                output_path,  # Update in place
                log
            )

            if not fva_result['success']:
                log.append(f"⚠ Warning: FvA update failed: {fva_result.get('error', 'Unknown error')}")
                log.append("  Continuing with other processing...")

        log.append(f"✓ Process completed successfully!")
        log.append("="*60)

        return {
            'success': True,
            'log': log,
            'output_file': output_filename,
            'stats': {
                'column': openpyxl.utils.get_column_letter(new_col),
                'total_balances': len(balances),
                'balances_matched': matched,
                'formulas_copied': formulas_copied + formulas_copied_148
            }
        }

    except Exception as e:
        log.append(f"\n✗ ERROR: {str(e)}")
        log.append(traceback.format_exc())
        return {
            'success': False,
            'log': log,
            'error': str(e)
        }


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_files():
    try:
        # Check if required files are present
        if 'weekly_file' not in request.files or 'rollforward_file' not in request.files:
            return jsonify({'success': False, 'error': 'Weekly Balances and Activity Rollforward files are required'})

        weekly_file = request.files['weekly_file']
        rollforward_file = request.files['rollforward_file']
        bth_file = request.files.get('bth_file')  # Optional
        aggregator_file = request.files.get('aggregator_file')  # Optional
        fva_1week = request.files.get('fva_1week')  # Optional
        fva_4week = request.files.get('fva_4week')  # Optional
        fva_13week = request.files.get('fva_13week')  # Optional

        # Check if required files are selected
        if weekly_file.filename == '' or rollforward_file.filename == '':
            return jsonify({'success': False, 'error': 'Please select Weekly Balances and Activity Rollforward files'})

        # Validate file types
        if not (allowed_file(weekly_file.filename) and allowed_file(rollforward_file.filename)):
            return jsonify({'success': False, 'error': 'Only .xlsx files are allowed'})

        if bth_file and bth_file.filename != '' and not allowed_file(bth_file.filename):
            return jsonify({'success': False, 'error': 'Only .xlsx files are allowed for BTH file'})

        if aggregator_file and aggregator_file.filename != '' and not allowed_file(aggregator_file.filename):
            return jsonify({'success': False, 'error': 'Only .xlsx files are allowed for Activity Aggregator file'})

        if fva_1week and fva_1week.filename != '' and not allowed_file(fva_1week.filename):
            return jsonify({'success': False, 'error': 'Only .xlsx files are allowed for 1-Week FvA file'})

        if fva_4week and fva_4week.filename != '' and not allowed_file(fva_4week.filename):
            return jsonify({'success': False, 'error': 'Only .xlsx files are allowed for 4-Week FvA file'})

        if fva_13week and fva_13week.filename != '' and not allowed_file(fva_13week.filename):
            return jsonify({'success': False, 'error': 'Only .xlsx files are allowed for 13-Week FvA file'})

        # Save uploaded files
        weekly_filename = secure_filename(weekly_file.filename)
        rollforward_filename = secure_filename(rollforward_file.filename)

        weekly_path = os.path.join(app.config['UPLOAD_FOLDER'], weekly_filename)
        rollforward_path = os.path.join(app.config['UPLOAD_FOLDER'], rollforward_filename)

        weekly_file.save(weekly_path)
        rollforward_file.save(rollforward_path)

        # Save BTH file if provided
        bth_path = None
        if bth_file and bth_file.filename != '':
            bth_filename = secure_filename(bth_file.filename)
            bth_path = os.path.join(app.config['UPLOAD_FOLDER'], bth_filename)
            bth_file.save(bth_path)

        # Save Activity Aggregator file if provided
        aggregator_path = None
        if aggregator_file and aggregator_file.filename != '':
            aggregator_filename = secure_filename(aggregator_file.filename)
            aggregator_path = os.path.join(app.config['UPLOAD_FOLDER'], aggregator_filename)
            aggregator_file.save(aggregator_path)

        # Save FvA files if provided
        fva_paths = {}
        if fva_1week and fva_1week.filename != '':
            fva_1week_filename = secure_filename(fva_1week.filename)
            fva_1week_path = os.path.join(app.config['UPLOAD_FOLDER'], fva_1week_filename)
            fva_1week.save(fva_1week_path)
            fva_paths['1week'] = fva_1week_path

        if fva_4week and fva_4week.filename != '':
            fva_4week_filename = secure_filename(fva_4week.filename)
            fva_4week_path = os.path.join(app.config['UPLOAD_FOLDER'], fva_4week_filename)
            fva_4week.save(fva_4week_path)
            fva_paths['4week'] = fva_4week_path

        if fva_13week and fva_13week.filename != '':
            fva_13week_filename = secure_filename(fva_13week.filename)
            fva_13week_path = os.path.join(app.config['UPLOAD_FOLDER'], fva_13week_filename)
            fva_13week.save(fva_13week_path)
            fva_paths['13week'] = fva_13week_path

        # Process the files
        result = process_files(weekly_path, rollforward_path, bth_path, aggregator_path, fva_paths if fva_paths else None)

        # Clean up uploaded files
        os.remove(weekly_path)
        os.remove(rollforward_path)
        if bth_path and os.path.exists(bth_path):
            os.remove(bth_path)
        if aggregator_path and os.path.exists(aggregator_path):
            os.remove(aggregator_path)
        for fva_path in fva_paths.values():
            if os.path.exists(fva_path):
                os.remove(fva_path)

        return jsonify(result)

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'log': [traceback.format_exc()]
        })


@app.route('/download/<filename>')
def download_file(filename):
    try:
        file_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
        return send_file(file_path, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 404


@app.route('/outputs')
def list_outputs():
    """List all output files"""
    try:
        files = []
        for filename in os.listdir(app.config['OUTPUT_FOLDER']):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
                files.append({
                    'name': filename,
                    'size': os.path.getsize(file_path),
                    'modified': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
                })
        files.sort(key=lambda x: x['modified'], reverse=True)
        return jsonify(files)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    # Create folders if they don't exist
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

    print("\n" + "="*60)
    print("Balance Updater Web App")
    print("="*60)
    print("\nStarting server...")
    print("Open your browser and go to: http://localhost:5000")
    print("\nPress CTRL+C to stop the server")
    print("="*60 + "\n")

    app.run(debug=True, host='0.0.0.0', port=8080)
