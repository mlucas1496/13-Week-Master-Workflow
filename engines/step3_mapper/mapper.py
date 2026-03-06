#!/usr/bin/env python3
"""
Activity Aggregator Mapper – Desktop Web App
=============================================
Run this script and a browser window opens. Drop in the weekly
Activity Aggregator and Activity Rollforward files, click "Map Activity",
and download the mapped output.

Usage:
    python app.py
"""

import os
import sys
import time
import uuid
import json
import threading
import webbrowser
import tempfile
from pathlib import Path
from collections import Counter

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from flask import Flask, render_template, request, jsonify, send_file

# ---------------------------------------------------------------------------
# Flask app setup
# ---------------------------------------------------------------------------
app = Flask(__name__)
APP_DIR = Path(__file__).parent
UPLOAD_DIR = Path(tempfile.mkdtemp(prefix="agg_mapper_"))
OUTPUT_DIR = Path(tempfile.mkdtemp(prefix="agg_mapper_out_"))

# Shared state for async processing
job_state = {
    "status": "idle",  # idle | running | done | error
    "logs": [],
    "progress": 0,
    "stats": {},
    "filename": "",
    "error": "",
}


def reset_state():
    job_state.update({
        "status": "idle",
        "logs": [],
        "progress": 0,
        "stats": {},
        "filename": "",
        "error": "",
    })


def add_log(text, cls=""):
    job_state["logs"].append({"text": text, "cls": cls})


# ---------------------------------------------------------------------------
# Mapping engine (from the proven script)
# ---------------------------------------------------------------------------

RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")

# ---------------------------------------------------------------------------
# Persistent Ref ID cache – saved between sessions
# ---------------------------------------------------------------------------
REF_ID_CACHE_PATH = APP_DIR / "ref_id_cache.json"


def load_ref_id_cache():
    """Load the saved Ref ID cache from disk."""
    if REF_ID_CACHE_PATH.exists():
        try:
            with open(REF_ID_CACHE_PATH, "r") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_ref_id_cache(cache):
    """Persist the Ref ID cache to disk for future weeks."""
    with open(REF_ID_CACHE_PATH, "w") as f:
        json.dump(cache, f, indent=2)


def build_ref_id_lookups(df_stacked):
    """
    Build Ref ID lookup tables from the Stacked Activity data.
    Returns a dict of lookup dicts, keyed by match strategy.

    Priority:
      1. Account Name + Notes (exact) → 13WCF Ref
      2. Account Name + Type + Sub Type → 13WCF Ref
      3. Account Name + Type → 13WCF Ref
      4. Account Name only → 13WCF Ref
    """
    ref_col = '13WCF Ref'

    # Only use rows where the ref is populated
    df_ref = df_stacked[df_stacked[ref_col].notna() & (df_stacked[ref_col] != 0)]

    # Tier 1: Account Name + Notes → most common Ref
    ref_by_notes = {}
    grp = df_ref[df_ref['Notes'].notna()].groupby(['Account Name', 'Notes'])[ref_col].agg(list)
    for (acct, notes), refs in grp.items():
        counter = Counter(refs)
        most_common, count = counter.most_common(1)[0]
        ref_by_notes[(str(acct), str(notes))] = int(most_common)

    # Tier 2: Account Name + Type + Sub Type
    ref_by_type_sub = {}
    grp = df_ref.groupby(['Account Name', 'Type', 'Sub Type'])[ref_col].agg(list)
    for (acct, typ, sub), refs in grp.items():
        counter = Counter(refs)
        most_common, _ = counter.most_common(1)[0]
        ref_by_type_sub[(str(acct), str(typ), str(sub))] = int(most_common)

    # Tier 3: Account Name + Type
    ref_by_type = {}
    grp = df_ref.groupby(['Account Name', 'Type'])[ref_col].agg(list)
    for (acct, typ), refs in grp.items():
        counter = Counter(refs)
        most_common, _ = counter.most_common(1)[0]
        ref_by_type[(str(acct), str(typ))] = int(most_common)

    # Tier 4: Account Name only
    ref_by_acct = {}
    grp = df_ref.groupby('Account Name')[ref_col].agg(list)
    for acct, refs in grp.items():
        counter = Counter(refs)
        most_common, _ = counter.most_common(1)[0]
        ref_by_acct[str(acct)] = int(most_common)

    return ref_by_notes, ref_by_type_sub, ref_by_type, ref_by_acct


def lookup_ref_id(row, ref_by_notes, ref_by_type_sub, ref_by_type, ref_by_acct, ref_cache):
    """
    Look up the 13WCF Ref # for a row missing it.
    Returns (ref_value, source) or (None, None) if not found.
    """
    acct = str(row['Account Name'])
    notes = str(row['Notes']) if pd.notna(row['Notes']) else None
    typ = str(row['Type']) if pd.notna(row['Type']) else None
    subtyp = str(row['Sub Type']) if pd.notna(row.get('Sub Type')) else None

    # Check persistent cache first
    cache_key = f"{acct}|{notes}|{typ}|{subtyp}"
    if cache_key in ref_cache:
        return ref_cache[cache_key], 'cache'

    # Tier 1: Account + Notes
    if notes and (acct, notes) in ref_by_notes:
        val = ref_by_notes[(acct, notes)]
        ref_cache[cache_key] = val
        return val, 'stacked_notes'

    # Tier 2: Account + Type + Sub Type
    if typ and subtyp and (acct, typ, subtyp) in ref_by_type_sub:
        val = ref_by_type_sub[(acct, typ, subtyp)]
        ref_cache[cache_key] = val
        return val, 'stacked_type_subtype'

    # Tier 3: Account + Type
    if typ and (acct, typ) in ref_by_type:
        val = ref_by_type[(acct, typ)]
        ref_cache[cache_key] = val
        return val, 'stacked_type'

    # Tier 4: Account only
    if acct in ref_by_acct:
        val = ref_by_acct[acct]
        ref_cache[cache_key] = val
        return val, 'stacked_acct'

    return None, None


def classify_direction(typ):
    typ_str = str(typ).lower()
    if any(kw in typ_str for kw in ['deposit', 'income', 'buy', 'credit', 'incoming', 'inflow']):
        return 'Inflow'
    elif any(kw in typ_str for kw in ['withdrawal', 'expense', 'sell', 'debit', 'outgoing', 'outflow']):
        return 'Outflow'
    return None


def get_rd_direction(rd_val):
    if pd.isna(rd_val):
        return None
    rd_str = str(rd_val).lower()
    if 'receipt' in rd_str:
        return 'Inflow'
    elif 'disbursement' in rd_str:
        return 'Outflow'
    return None


HSBC_NOTES_RULES = {
    'AUTOPAY OUT': 'All Other',
    'ADP': 'Staff & Bonus',
    'BULLISH HK LIMITED': 'Intercompany Inflow',
    'NONREF': 'Intercompany Outflow',
    'INTEREST': 'Rewards and Other Interest',
    'CREDIT INTEREST': 'Rewards and Other Interest',
    'ELECTRIC': 'All Other',
    'HKT': 'All Other',
    'CSL MOBILE': 'All Other',
    'HSBCNET': 'All Other',
    'FEE': 'All Other',
    'VODAFONE': 'All Other',
}

BC_NOTES_RULES = {
    'BULLISH GLOBAL': 'Intercompany Inflow',
    'BULLISH EUROPE': 'Intercompany Outflow',
    'INTERNAL TRANSFER': 'Intercompany Outflow',
    'SOCIETE GENERALE': 'Rewards and Other Interest',
    'ALLUNITY': 'Rewards and Other Interest',
    'ADP EMPLOYER': 'Consulting, Contracting, Prof Fees',
    'PRICEWATERHOUSECOOPERS': 'Consulting, Contracting, Prof Fees',
    'EXPENSES REIMBURSEMENTS': 'Consulting, Contracting, Prof Fees',
    'RECHTSANWALTE': 'All Other',
    'OFFICE GROUP': 'Facilities Rent',
    'BCB PAYMENTS': 'Consulting, Contracting, Prof Fees',
    'CENTRICCOM': 'Consulting, Contracting, Prof Fees',
    'BUNDESANZEIGER': 'All Other',
}


def build_lookups(df_stacked):
    """Build the tiered mapping lookup tables from the Stacked Activity data."""

    # TIER 1: Account Name + Notes
    notes_mapping = {}
    df_notes = df_stacked[df_stacked['Notes'].notna()].groupby(
        ['Account Name', 'Notes'])['13WCF Line'].agg(list)
    for (acct, notes), lines in df_notes.items():
        counter = Counter(lines)
        most_common, count = counter.most_common(1)[0]
        total = len(lines)
        notes_mapping[(acct, notes)] = (most_common, count / total, total)

    # TIER 2: Account Name + Type + Sub Type
    type_subtype_mapping = {}
    df_ts = df_stacked.groupby(['Account Name', 'Type', 'Sub Type'])['13WCF Line'].agg(list)
    for (acct, typ, subtyp), lines in df_ts.items():
        counter = Counter(lines)
        most_common, count = counter.most_common(1)[0]
        total = len(lines)
        type_subtype_mapping[(acct, typ, subtyp)] = (most_common, count / total, total)

    # TIER 3: Account Name + Type
    type_mapping = {}
    df_t = df_stacked.groupby(['Account Name', 'Type'])['13WCF Line'].agg(list)
    for (acct, typ), lines in df_t.items():
        counter = Counter(lines)
        most_common, count = counter.most_common(1)[0]
        total = len(lines)
        type_mapping[(acct, typ)] = (most_common, count / total, total)

    # TIER 4: Account Name + Direction
    direction_mapping = {}
    df_stacked['_direction'] = df_stacked['Type'].apply(classify_direction)
    df_dir = df_stacked[df_stacked['_direction'].notna()].groupby(
        ['Account Name', '_direction'])['13WCF Line'].agg(list)
    for (acct, direction), lines in df_dir.items():
        counter = Counter(lines)
        most_common, count = counter.most_common(1)[0]
        total = len(lines)
        direction_mapping[(acct, direction)] = (most_common, count / total, total)

    # TIER 5: Account Name only
    acct_mapping = {}
    df_a = df_stacked.groupby('Account Name')['13WCF Line'].agg(list)
    for acct, lines in df_a.items():
        counter = Counter(lines)
        most_common, count = counter.most_common(1)[0]
        total = len(lines)
        acct_mapping[acct] = (most_common, count / total, total)

    return notes_mapping, type_subtype_mapping, type_mapping, direction_mapping, acct_mapping


def determine_mapping(row, notes_mapping, type_subtype_mapping, type_mapping,
                      direction_mapping, acct_mapping):
    """Determine the 13WCF Line Item Mapping for an unmapped row."""
    acct = row['Account Name']
    acct_str = str(acct)
    notes = row['Notes'] if pd.notna(row['Notes']) else None
    notes_upper = str(notes).upper() if notes and pd.notna(notes) else ''
    typ = row['Type']
    typ_str = str(typ).lower() if pd.notna(typ) else ''
    subtyp = row['Sub Type'] if pd.notna(row.get('Sub Type')) else None
    rd = row.get('R/D')
    direction = get_rd_direction(rd) if rd is not None else None

    # =========================================================================
    # HARD RULES: Known account mappings (always applied first)
    # =========================================================================

    # Airtable → always RCFs
    if 'Airtable' in acct_str:
        return ('RCFs', 0.99, 'HardRule_Airtable')

    # Atlantic Union MMA (LOC Collateral) - Bullish US
    # Atlantic Union MMA 048 - Bullish US
    # → Intercompany Inflow, Intercompany Outflow, or Rewards and Other Interest
    if 'Atlantic Union' in acct_str:
        if 'interest' in typ_str or 'interest' in notes_upper.lower():
            return ('Rewards and Other Interest', 0.99, 'HardRule_AtlanticUnion')
        if direction == 'Inflow':
            return ('Intercompany Inflow', 0.99, 'HardRule_AtlanticUnion')
        elif direction == 'Outflow':
            return ('Intercompany Outflow', 0.99, 'HardRule_AtlanticUnion')
        return ('Intercompany Inflow', 0.90, 'HardRule_AtlanticUnion')

    # Any Binance - BTH account → Intercompany Inflow or Outflow
    if 'Binance' in acct_str and 'BTH' in acct_str:
        if direction == 'Inflow':
            return ('Intercompany Inflow', 0.99, 'HardRule_BinanceBTH')
        elif direction == 'Outflow':
            return ('Intercompany Outflow', 0.99, 'HardRule_BinanceBTH')
        return ('Intercompany Inflow', 0.90, 'HardRule_BinanceBTH')

    # BCM - Fireblocks → Intercompany Inflow or Outflow
    if 'BCM' in acct_str and 'Fireblocks' in acct_str:
        if direction == 'Inflow':
            return ('Intercompany Inflow', 0.99, 'HardRule_BCMFireblocks')
        elif direction == 'Outflow':
            return ('Intercompany Outflow', 0.99, 'HardRule_BCMFireblocks')
        return ('Intercompany Inflow', 0.90, 'HardRule_BCMFireblocks')

    # =========================================================================
    # SPECIAL RULES (pattern-based)
    # =========================================================================

    # SPECIAL RULES: HSBC
    if 'HSBC' in str(acct):
        if notes and pd.notna(notes):
            notes_upper = str(notes).upper()
            for pattern, mapping in HSBC_NOTES_RULES.items():
                if pattern in notes_upper:
                    return (mapping, 0.85, 'HSBC_Notes_Rule')
        if direction == 'Inflow':
            return ('All Other', 0.35, 'HSBC_Default')
        elif direction == 'Outflow':
            return ('All Other', 0.35, 'HSBC_Default')

    # SPECIAL RULES: Banking Circle
    if 'Banking Circle' in str(acct):
        if notes and pd.notna(notes):
            notes_upper = str(notes).upper()
            for pattern, mapping in BC_NOTES_RULES.items():
                if pattern in notes_upper:
                    return (mapping, 0.80, 'BC_Notes_Rule')
        if acct in acct_mapping:
            m, c, n = acct_mapping[acct]
            return (m, c * 0.5, 'BC_Account_Default')

    # SPECIAL RULES: FTD
    if 'FTD' in str(acct):
        if direction == 'Inflow':
            return ('Intercompany Inflow', 0.85, 'FTD_Direction')
        elif direction == 'Outflow':
            return ('Intercompany Outflow', 0.85, 'FTD_Direction')

    # TIER 1: Account Name + Notes
    if notes and pd.notna(notes) and (acct, notes) in notes_mapping:
        m, c, n = notes_mapping[(acct, notes)]
        adj = min(c, 0.95) if n >= 3 else c * 0.7
        return (m, adj, 'Tier1_Notes')

    # TIER 2: Account Name + Type + Sub Type
    if subtyp and (acct, typ, subtyp) in type_subtype_mapping:
        m, c, n = type_subtype_mapping[(acct, typ, subtyp)]
        adj = min(c, 0.95) if n >= 5 else c * 0.8
        return (m, adj, 'Tier2_TypeSubType')

    # TIER 3: Account Name + Type
    if (acct, typ) in type_mapping:
        m, c, n = type_mapping[(acct, typ)]
        adj = c * 0.85 if n >= 5 else c * 0.6
        return (m, adj, 'Tier3_Type')

    # TIER 4: Account Name + Direction
    if direction and (acct, direction) in direction_mapping:
        m, c, n = direction_mapping[(acct, direction)]
        adj = c * 0.7 if n >= 5 else c * 0.5
        return (m, adj, 'Tier4_Direction')

    # TIER 5: Account Name only
    if acct in acct_mapping:
        m, c, n = acct_mapping[acct]
        return (m, c * 0.5, 'Tier5_Account')

    # FALLBACK: Crypto accounts
    if any(kw in str(acct) for kw in ['Binance', 'Coinbase', 'CB_']):
        if direction == 'Inflow':
            return ('Intercompany Inflow', 0.50, 'Fallback_Crypto')
        elif direction == 'Outflow':
            return ('Intercompany Outflow', 0.50, 'Fallback_Crypto')

    # LAST RESORT
    if direction == 'Inflow':
        return ('All Other', 0.15, 'LastResort')
    elif direction == 'Outflow':
        return ('All Other', 0.15, 'LastResort')
    return ('All Other', 0.10, 'NoMatch')


def run_mapping(agg_path, roll_path):
    """Main mapping function – runs in a background thread."""
    try:
        reset_state()
        job_state["status"] = "running"

        # ---- Load Rollforward ----
        add_log("Loading Activity Rollforward (Stacked Activity)...", "step")
        job_state["progress"] = 10
        df_stacked = pd.read_excel(roll_path, sheet_name="Stacked Activity", engine="openpyxl")
        add_log(f"  Loaded {len(df_stacked):,} rows from Stacked Activity", "")
        job_state["progress"] = 25

        # ---- Build lookups ----
        add_log("Building mapping lookup tables...", "step")
        lookups = build_lookups(df_stacked)
        notes_mapping, type_subtype_mapping, type_mapping, direction_mapping, acct_mapping = lookups
        add_log(f"  Built {len(notes_mapping):,} Account+Notes combos (Tier 1)", "")
        add_log(f"  Built {len(type_subtype_mapping):,} Account+Type+SubType combos (Tier 2)", "")
        add_log(f"  Built {len(type_mapping):,} Account+Type combos (Tier 3)", "")

        # ---- Build Ref ID lookups ----
        add_log("Building Ref ID lookup tables...", "step")
        ref_by_notes, ref_by_type_sub, ref_by_type, ref_by_acct = build_ref_id_lookups(df_stacked)
        ref_cache = load_ref_id_cache()
        add_log(f"  Ref ID lookups: {len(ref_by_notes):,} by Notes, {len(ref_by_type):,} by Type, {len(ref_by_acct):,} by Account", "")
        add_log(f"  Loaded {len(ref_cache):,} cached Ref IDs from previous runs", "")
        job_state["progress"] = 35

        # ---- Load Aggregator with pandas ----
        add_log("Loading Activity Aggregator (Alteryx_Output)...", "step")
        df_agg = pd.read_excel(agg_path, sheet_name="Alteryx_Output", engine="openpyxl")
        add_log(f"  Loaded {len(df_agg):,} rows from Alteryx_Output", "")
        job_state["progress"] = 50

        # ---- Find unmapped rows ----
        unmapped_mask = (
            (df_agg['Manual User Check'] == 'Not Mapped - Check') &
            (df_agg['13WCF Line Item Mapping'].isna() | (df_agg['13WCF Line Item Mapping'] == ''))
        )
        unmapped_indices = df_agg[unmapped_mask].index.tolist()
        add_log(f"  Found {len(unmapped_indices):,} unmapped rows", "warn")
        job_state["progress"] = 55

        if len(unmapped_indices) == 0:
            add_log("No unmapped rows found. Nothing to do.", "success")
            job_state["status"] = "done"
            job_state["progress"] = 100
            job_state["stats"] = {"high": 0, "medium": 0, "low": 0}
            # Still save a copy
            out_name = "Activity Aggregator - MAPPED.xlsx"
            out_path = OUTPUT_DIR / out_name
            import shutil
            shutil.copy2(agg_path, out_path)
            job_state["filename"] = out_name
            return

        # ---- Determine mappings ----
        add_log("Determining mappings for each unmapped row...", "step")
        results = []
        confidence_buckets = {'high': 0, 'medium': 0, 'low': 0}
        tier_counts = Counter()

        for i, idx in enumerate(unmapped_indices):
            row = df_agg.iloc[idx]
            mapping, confidence, tier = determine_mapping(
                row, notes_mapping, type_subtype_mapping, type_mapping,
                direction_mapping, acct_mapping
            )
            results.append((idx, mapping, confidence, tier))
            tier_counts[tier] += 1
            if confidence >= 0.7:
                confidence_buckets['high'] += 1
            elif confidence >= 0.4:
                confidence_buckets['medium'] += 1
            else:
                confidence_buckets['low'] += 1

            if (i + 1) % 2000 == 0:
                pct = 55 + int(20 * (i + 1) / len(unmapped_indices))
                job_state["progress"] = pct
                add_log(f"  Processed {i+1:,} / {len(unmapped_indices):,} rows...", "")

        add_log(f"  High confidence:   {confidence_buckets['high']:,} rows", "success")
        add_log(f"  Medium confidence:  {confidence_buckets['medium']:,} rows (yellow)", "warn")
        add_log(f"  Low confidence:     {confidence_buckets['low']:,} rows (red)", "error" if confidence_buckets['low'] > 0 else "")

        for tier, count in sorted(tier_counts.items(), key=lambda x: -x[1]):
            add_log(f"    {tier}: {count:,}", "")

        job_state["progress"] = 75

        # ---- Write to Excel with openpyxl ----
        add_log("Loading workbook for writing (this may take a minute)...", "step")
        wb = openpyxl.load_workbook(agg_path)
        ws = wb["Alteryx_Output"]
        job_state["progress"] = 85

        COL_X = 24   # 13WCF Line Item Mapping
        mapped_count = 0

        add_log("Applying mappings and highlighting...", "step")
        for idx, mapping, confidence, tier in results:
            excel_row = idx + 2
            cell_x = ws.cell(row=excel_row, column=COL_X)
            cell_x.value = mapping
            mapped_count += 1

            if confidence < 0.4:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=excel_row, column=col).fill = RED_FILL
            elif confidence < 0.7:
                cell_x.fill = YELLOW_FILL

        job_state["progress"] = 88

        # ---- Ref ID fill: find and fix missing 13WCF Ref # (column K = 11) ----
        add_log("Scanning for missing Ref IDs (13WCF Ref #)...", "step")
        COL_K = 11   # 13WCF Ref #
        ref_col_name = '13WCF Ref #'

        # Column AC contains mapping status (Mapped / Unmapped / Excluded)
        col_ac_name = df_agg.columns[28]  # Column AC = index 28

        # Only consider rows that are Mapped or Unmapped (skip Excluded)
        not_excluded_mask = ~df_agg[col_ac_name].astype(str).str.lower().str.contains('exclud', na=False)

        # Find rows with blank/zero Ref # that are NOT excluded
        missing_ref_mask = (
            (
                df_agg[ref_col_name].isna() |
                (df_agg[ref_col_name] == '') |
                (df_agg[ref_col_name] == 0)
            ) & not_excluded_mask
        )
        missing_ref_indices = df_agg[missing_ref_mask].index.tolist()
        add_log(f"  Found {len(missing_ref_indices):,} non-excluded rows with missing/zero Ref ID", "warn")

        ref_filled = 0
        ref_not_found = 0

        for idx in missing_ref_indices:
            row = df_agg.iloc[idx]
            excel_row = idx + 2
            ref_val, source = lookup_ref_id(
                row, ref_by_notes, ref_by_type_sub, ref_by_type, ref_by_acct, ref_cache
            )
            cell_k = ws.cell(row=excel_row, column=COL_K)

            if ref_val is not None:
                # Found a Ref ID – fill it in and highlight orange
                cell_k.value = ref_val
                cell_k.fill = ORANGE_FILL
                ref_filled += 1
            else:
                # Could not find – highlight orange to flag it
                cell_k.fill = ORANGE_FILL
                ref_not_found += 1

        add_log(f"  Filled {ref_filled:,} Ref IDs from Stacked Activity / cache (orange)", "success")
        if ref_not_found > 0:
            add_log(f"  Could not find Ref ID for {ref_not_found:,} rows (orange, empty)", "warn")

        # Save the Ref ID cache for future weeks
        save_ref_id_cache(ref_cache)
        add_log(f"  Saved {len(ref_cache):,} Ref IDs to cache for future use", "")

        job_state["progress"] = 94

        # ---- Save output ----
        out_name = "Activity Aggregator - MAPPED.xlsx"
        out_path = OUTPUT_DIR / out_name
        add_log(f"Saving mapped file...", "step")
        wb.save(str(out_path))
        add_log(f"Saved! {mapped_count:,} rows mapped, {ref_filled:,} Ref IDs filled.", "success")

        job_state["progress"] = 100
        job_state["stats"] = {
            **confidence_buckets,
            "ref_filled": ref_filled,
            "ref_not_found": ref_not_found,
        }
        job_state["filename"] = out_name
        job_state["status"] = "done"

    except Exception as e:
        import traceback
        add_log(f"ERROR: {str(e)}", "error")
        add_log(traceback.format_exc(), "error")
        job_state["status"] = "error"
        job_state["error"] = str(e)


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/process", methods=["POST"])
def process():
    if job_state["status"] == "running":
        return jsonify({"error": "A job is already running"}), 409

    agg_file = request.files.get("aggregator")
    roll_file = request.files.get("rollforward")

    if not agg_file or not roll_file:
        return jsonify({"error": "Both files are required"}), 400

    # Save uploaded files
    agg_path = UPLOAD_DIR / f"agg_{uuid.uuid4().hex[:8]}.xlsx"
    roll_path = UPLOAD_DIR / f"roll_{uuid.uuid4().hex[:8]}.xlsx"
    agg_file.save(str(agg_path))
    roll_file.save(str(roll_path))

    # Run mapping in background thread
    thread = threading.Thread(target=run_mapping, args=(str(agg_path), str(roll_path)), daemon=True)
    thread.start()

    return jsonify({"status": "started"})


@app.route("/status")
def status():
    return jsonify({
        "status": job_state["status"],
        "logs": job_state["logs"],
        "progress": job_state["progress"],
        "stats": job_state["stats"],
        "filename": job_state["filename"],
        "error": job_state["error"],
    })


@app.route("/download/<filename>")
def download(filename):
    path = OUTPUT_DIR / filename
    if not path.exists():
        return jsonify({"error": "File not found"}), 404
    return send_file(
        str(path),
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def open_browser():
    """Open browser after a short delay to let Flask start."""
    time.sleep(1.5)
    webbrowser.open("http://127.0.0.1:5051")


if __name__ == "__main__":
    print("=" * 60)
    print("  Activity Aggregator Mapper")
    print("  Opening in your browser at http://127.0.0.1:5051")
    print("  Press Ctrl+C to quit")
    print("=" * 60)

    threading.Thread(target=open_browser, daemon=True).start()
    app.run(host="127.0.0.1", port=5051, debug=False)
