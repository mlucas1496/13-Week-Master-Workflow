"""
13WCF Unified Workflow
======================
Combines all 5 weekly processing steps into a single Flask application:
  Step 1: Weekly Balances Rollforward (in-browser)
  Step 2: Activity Aggregator Update
  Step 3: Activity Aggregator Mapper
  Step 4: Activity Rollforward
  Step 5: 13WCF Data Loader (in-browser)

All engine code is bundled in the engines/ directory — no external
folders or files are needed.
"""

import os
import sys
import uuid
import json
import time
import shutil
import threading
import traceback
import importlib.util
from datetime import datetime

from flask import Flask, render_template, request, jsonify, Response, send_file
from werkzeug.utils import secure_filename

# ---------------------------------------------------------------------------
# Path setup: all engines are bundled locally
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

STEP2_ENGINE_DIR = os.path.join(BASE_DIR, "engines", "step2_aggregator")
STEP3_ENGINE_DIR = os.path.join(BASE_DIR, "engines", "step3_mapper")
STEP4_ENGINE_DIR = os.path.join(BASE_DIR, "engines", "step4_rollforward")
STEP5_ENGINE_DIR = os.path.join(BASE_DIR, "engines", "step5_data_loader")

# Step 2 uses `from pipeline.xxx import ...` and `from config import ...`
# so we add its directory to sys.path
if STEP2_ENGINE_DIR not in sys.path:
    sys.path.insert(0, STEP2_ENGINE_DIR)

# Step 4 uses `import stacked_activity_updater` and `import fva_data_updater`
# as sibling imports, so we add its directory too
if STEP4_ENGINE_DIR not in sys.path:
    sys.path.insert(0, STEP4_ENGINE_DIR)

# ---------------------------------------------------------------------------
# Import Step 2 pipeline (Activity Aggregator Update)
# ---------------------------------------------------------------------------
from pipeline.orchestrator import run_pipeline as step2_run_pipeline

# ---------------------------------------------------------------------------
# Import Step 3 mapper (Activity Aggregator Mapper)
# Uses importlib to avoid Flask app conflicts
# ---------------------------------------------------------------------------
_mapper_spec = importlib.util.spec_from_file_location(
    "mapper_engine",
    os.path.join(STEP3_ENGINE_DIR, "mapper.py"),
)
mapper_engine = importlib.util.module_from_spec(_mapper_spec)
_orig_argv = sys.argv
sys.argv = [""]
_mapper_spec.loader.exec_module(mapper_engine)
sys.argv = _orig_argv

# ---------------------------------------------------------------------------
# Import Step 4 modules (Activity Rollforward)
# ---------------------------------------------------------------------------
import stacked_activity_updater
import fva_data_updater

_rollforward_spec = importlib.util.spec_from_file_location(
    "rollforward_engine",
    os.path.join(STEP4_ENGINE_DIR, "rollforward.py"),
)
rollforward_engine = importlib.util.module_from_spec(_rollforward_spec)
sys.argv = [""]
_rollforward_spec.loader.exec_module(rollforward_engine)
sys.argv = _orig_argv

# ---------------------------------------------------------------------------
# Flask app
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 1024 * 1024 * 1024  # 1 GB
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Configure the imported rollforward engine's app to use our folders
rollforward_engine.app.config["OUTPUT_FOLDER"] = OUTPUT_DIR
rollforward_engine.app.config["UPLOAD_FOLDER"] = UPLOAD_DIR

# ---------------------------------------------------------------------------
# Session state
# ---------------------------------------------------------------------------
workflow = {
    "files": {},          # key -> {"path": ..., "name": ...}
    "step_status": {      # step_id -> "pending" | "running" | "done" | "error"
        "1": "pending",
        "2": "pending",
        "3": "pending",
        "4": "pending",
        "5": "pending",
    },
    "step_outputs": {},   # step_id -> {"path": ..., "name": ...}
    "step_logs": {},      # step_id -> [log_messages]
    "step_stages": {},    # step_id -> {stage_id: status}
    "step_errors": {},    # step_id -> error_message
    "step_stats": {},     # step_id -> stats dict
}

ALLOWED_EXT = {"xlsx", "xls", "csv"}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXT


def save_upload(file_obj, key):
    """Save an uploaded file and register it in the workflow."""
    fname = secure_filename(file_obj.filename)
    unique = uuid.uuid4().hex[:8]
    dest = os.path.join(UPLOAD_DIR, f"{unique}_{fname}")
    file_obj.save(dest)
    workflow["files"][key] = {"path": dest, "name": file_obj.filename}
    return dest


def get_file_path(key):
    """Get the path for a registered file (uploaded or output from prior step)."""
    entry = workflow["files"].get(key)
    return entry["path"] if entry else None


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/status")
def api_status():
    """Return full workflow state."""
    return jsonify({
        "files": {k: v["name"] for k, v in workflow["files"].items()},
        "step_status": workflow["step_status"],
        "step_outputs": {
            k: v["name"] for k, v in workflow["step_outputs"].items()
        },
        "step_stats": workflow["step_stats"],
    })


@app.route("/api/upload", methods=["POST"])
def api_upload():
    """Upload one or more files. Form keys become file registry keys."""
    saved = {}
    for key in request.files:
        f = request.files[key]
        if f and f.filename:
            path = save_upload(f, key)
            saved[key] = f.filename
    return jsonify({"saved": saved})


@app.route("/api/save-output/<step>", methods=["POST"])
def api_save_output(step):
    """Save a browser-processed output file (Steps 1 & 5)."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400
    fname = secure_filename(f.filename)
    unique = uuid.uuid4().hex[:8]
    dest = os.path.join(OUTPUT_DIR, f"step{step}_{unique}_{fname}")
    f.save(dest)
    workflow["step_outputs"][step] = {"path": dest, "name": f.filename}
    workflow["step_status"][step] = "done"

    # Register output as available for subsequent steps
    if step == "1":
        workflow["files"]["step1_weekly_balances_output"] = {
            "path": dest, "name": f.filename
        }
    elif step == "5":
        workflow["files"]["step5_13wcf_output"] = {
            "path": dest, "name": f.filename
        }

    return jsonify({"saved": fname})


# ---------------------------------------------------------------------------
# Step 2: Activity Aggregator Update
# ---------------------------------------------------------------------------

@app.route("/api/run/2", methods=["POST"])
def run_step2():
    if workflow["step_status"]["2"] == "running":
        return jsonify({"error": "Already running"}), 409

    # Check required files
    required = ["s2_prev_week", "s2_bank_statements", "s2_all_transactions",
                 "s2_loan_report", "s2_search_strings", "s2_static_mapping"]
    missing = [k for k in required if not get_file_path(k)]
    if missing:
        return jsonify({"error": f"Missing files: {', '.join(missing)}"}), 400

    workflow["step_status"]["2"] = "running"
    workflow["step_logs"]["2"] = []
    workflow["step_stages"]["2"] = {}
    workflow["step_errors"]["2"] = None

    file_paths = {
        "prev_week": get_file_path("s2_prev_week"),
        "bank_statements": get_file_path("s2_bank_statements"),
        "all_transactions": get_file_path("s2_all_transactions"),
        "loan_report": get_file_path("s2_loan_report"),
        "search_strings": get_file_path("s2_search_strings"),
        "static_mapping": get_file_path("s2_static_mapping"),
    }

    def worker():
        try:
            output_dir = os.path.join(OUTPUT_DIR, "step2")
            os.makedirs(output_dir, exist_ok=True)

            def log(msg):
                workflow["step_logs"]["2"].append(msg)

            def set_stage(stage_id, status):
                workflow["step_stages"]["2"][stage_id] = status

            result = step2_run_pipeline(file_paths, output_dir, log, set_stage)

            workflow["step_outputs"]["2"] = {
                "path": result["file_path"],
                "name": result["file_name"],
            }
            workflow["step_stats"]["2"] = result.get("stats", {})
            workflow["step_status"]["2"] = "done"

            # Register output for Step 3
            workflow["files"]["step2_activity_aggregator_output"] = {
                "path": result["file_path"],
                "name": result["file_name"],
            }

        except Exception as e:
            workflow["step_errors"]["2"] = str(e)
            workflow["step_status"]["2"] = "error"
            workflow["step_logs"]["2"].append(f"ERROR: {e}")
            traceback.print_exc()

    threading.Thread(target=worker, daemon=True).start()
    return jsonify({"status": "started"})


@app.route("/api/progress/2")
def progress_step2():
    """SSE endpoint for Step 2 progress."""
    def generate():
        last_log_idx = 0
        last_stages = {}
        while True:
            logs = workflow["step_logs"].get("2", [])
            if len(logs) > last_log_idx:
                for msg in logs[last_log_idx:]:
                    yield f"data: {json.dumps({'type': 'log', 'message': msg})}\n\n"
                last_log_idx = len(logs)

            stages = dict(workflow["step_stages"].get("2", {}))
            if stages != last_stages:
                yield f"data: {json.dumps({'type': 'stages', 'stages': stages})}\n\n"
                last_stages = stages

            status = workflow["step_status"]["2"]
            if status == "done":
                stats = workflow["step_stats"].get("2", {})
                yield f"data: {json.dumps({'type': 'done', 'stats': stats})}\n\n"
                break
            elif status == "error":
                yield f"data: {json.dumps({'type': 'error', 'message': workflow['step_errors'].get('2', 'Unknown')})}\n\n"
                break
            time.sleep(0.3)

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ---------------------------------------------------------------------------
# Step 3: Activity Aggregator Mapper
# ---------------------------------------------------------------------------

@app.route("/api/run/3", methods=["POST"])
def run_step3():
    if workflow["step_status"]["3"] == "running":
        return jsonify({"error": "Already running"}), 409

    # Activity Aggregator: from Step 2 output or manual upload
    agg_path = get_file_path("step2_activity_aggregator_output") or get_file_path("s3_aggregator")
    roll_path = get_file_path("s3_rollforward")

    if not agg_path:
        return jsonify({"error": "Missing Activity Aggregator file (run Step 2 first or upload manually)"}), 400
    if not roll_path:
        return jsonify({"error": "Missing Activity Rollforward file"}), 400

    workflow["step_status"]["3"] = "running"
    workflow["step_logs"]["3"] = []
    workflow["step_errors"]["3"] = None

    def worker():
        try:
            mapper_engine.run_mapping(agg_path, roll_path)

            # Wait for completion
            while mapper_engine.job_state["status"] == "running":
                current_logs = mapper_engine.job_state.get("logs", [])
                for entry in current_logs[len(workflow["step_logs"]["3"]):]:
                    text = entry["text"] if isinstance(entry, dict) else str(entry)
                    workflow["step_logs"]["3"].append(text)
                time.sleep(0.5)

            # Final log sync
            current_logs = mapper_engine.job_state.get("logs", [])
            for entry in current_logs[len(workflow["step_logs"]["3"]):]:
                text = entry["text"] if isinstance(entry, dict) else str(entry)
                workflow["step_logs"]["3"].append(text)

            if mapper_engine.job_state["status"] == "done":
                src_name = mapper_engine.job_state.get("filename", "Activity Aggregator - MAPPED.xlsx")
                src_path = str(mapper_engine.OUTPUT_DIR / src_name)

                dest_name = f"Activity_Aggregator_MAPPED_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                dest_path = os.path.join(OUTPUT_DIR, dest_name)
                shutil.copy2(src_path, dest_path)

                workflow["step_outputs"]["3"] = {
                    "path": dest_path,
                    "name": dest_name,
                }
                workflow["step_stats"]["3"] = mapper_engine.job_state.get("stats", {})
                workflow["step_status"]["3"] = "done"

                # Register for Step 4
                workflow["files"]["step3_mapped_aggregator_output"] = {
                    "path": dest_path,
                    "name": dest_name,
                }
            else:
                workflow["step_errors"]["3"] = mapper_engine.job_state.get("error", "Unknown error")
                workflow["step_status"]["3"] = "error"

        except Exception as e:
            workflow["step_errors"]["3"] = str(e)
            workflow["step_status"]["3"] = "error"
            workflow["step_logs"]["3"].append(f"ERROR: {e}")
            traceback.print_exc()

    threading.Thread(target=worker, daemon=True).start()
    return jsonify({"status": "started"})


@app.route("/api/progress/3")
def progress_step3():
    """SSE endpoint for Step 3 progress."""
    def generate():
        last_log_idx = 0
        while True:
            logs = workflow["step_logs"].get("3", [])
            if len(logs) > last_log_idx:
                for msg in logs[last_log_idx:]:
                    yield f"data: {json.dumps({'type': 'log', 'message': msg})}\n\n"
                last_log_idx = len(logs)

            pct = mapper_engine.job_state.get("progress", 0)
            yield f"data: {json.dumps({'type': 'progress', 'percent': pct})}\n\n"

            status = workflow["step_status"]["3"]
            if status == "done":
                stats = workflow["step_stats"].get("3", {})
                yield f"data: {json.dumps({'type': 'done', 'stats': stats})}\n\n"
                break
            elif status == "error":
                yield f"data: {json.dumps({'type': 'error', 'message': workflow['step_errors'].get('3', 'Unknown')})}\n\n"
                break
            time.sleep(0.5)

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ---------------------------------------------------------------------------
# Step 4: Activity Rollforward
# ---------------------------------------------------------------------------

@app.route("/api/run/4", methods=["POST"])
def run_step4():
    if workflow["step_status"]["4"] == "running":
        return jsonify({"error": "Already running"}), 409

    weekly_path = get_file_path("step1_weekly_balances_output") or get_file_path("s4_weekly_balances")
    rollforward_path = get_file_path("s3_rollforward") or get_file_path("s4_rollforward")
    aggregator_path = get_file_path("step3_mapped_aggregator_output") or get_file_path("s4_aggregator")
    bth_path = get_file_path("s4_bth")
    fva_1w = get_file_path("s4_fva_1week")
    fva_4w = get_file_path("s4_fva_4week")
    fva_13w = get_file_path("s4_fva_13week")

    if not weekly_path:
        return jsonify({"error": "Missing Weekly Balances file (run Step 1 first or upload manually)"}), 400
    if not rollforward_path:
        return jsonify({"error": "Missing Activity Rollforward file"}), 400

    workflow["step_status"]["4"] = "running"
    workflow["step_logs"]["4"] = []
    workflow["step_errors"]["4"] = None

    def worker():
        try:
            fva_files = {}
            if fva_1w:
                fva_files["1week"] = fva_1w
            if fva_4w:
                fva_files["4week"] = fva_4w
            if fva_13w:
                fva_files["13week"] = fva_13w

            result = rollforward_engine.process_files(
                weekly_file_path=weekly_path,
                rollforward_file_path=rollforward_path,
                bth_file_path=bth_path,
                aggregator_file_path=aggregator_path,
                fva_files=fva_files if fva_files else None,
            )

            workflow["step_logs"]["4"] = result.get("log", [])

            if result.get("success"):
                output_name = result["output_file"]
                output_path = os.path.join(OUTPUT_DIR, output_name)

                workflow["step_outputs"]["4"] = {
                    "path": output_path,
                    "name": output_name,
                }
                workflow["step_stats"]["4"] = result.get("stats", {})
                workflow["step_status"]["4"] = "done"

                workflow["files"]["step4_rollforward_output"] = {
                    "path": output_path,
                    "name": output_name,
                }
            else:
                workflow["step_errors"]["4"] = result.get("error", "Processing failed")
                workflow["step_status"]["4"] = "error"

        except Exception as e:
            workflow["step_errors"]["4"] = str(e)
            workflow["step_status"]["4"] = "error"
            workflow["step_logs"]["4"].append(f"ERROR: {e}")
            traceback.print_exc()

    threading.Thread(target=worker, daemon=True).start()
    return jsonify({"status": "started"})


@app.route("/api/progress/4")
def progress_step4():
    """SSE endpoint for Step 4 progress."""
    def generate():
        last_log_idx = 0
        while True:
            logs = workflow["step_logs"].get("4", [])
            if len(logs) > last_log_idx:
                for msg in logs[last_log_idx:]:
                    yield f"data: {json.dumps({'type': 'log', 'message': msg})}\n\n"
                last_log_idx = len(logs)

            status = workflow["step_status"]["4"]
            if status == "done":
                stats = workflow["step_stats"].get("4", {})
                yield f"data: {json.dumps({'type': 'done', 'stats': stats})}\n\n"
                break
            elif status == "error":
                yield f"data: {json.dumps({'type': 'error', 'message': workflow['step_errors'].get('4', 'Unknown')})}\n\n"
                break
            time.sleep(0.5)

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ---------------------------------------------------------------------------
# Download any step output
# ---------------------------------------------------------------------------

@app.route("/api/download/<step>")
def download_output(step):
    output = workflow["step_outputs"].get(step)
    if not output:
        return jsonify({"error": "No output available"}), 404
    return send_file(
        output["path"],
        as_attachment=True,
        download_name=output["name"],
    )


# ---------------------------------------------------------------------------
# Step 1: Extract column K values server-side (openpyxl data_only=True)
# ---------------------------------------------------------------------------

@app.route("/api/extract-k-values", methods=["POST"])
def extract_k_values():
    """
    Read column K values from the Weekly Balances file by evaluating
    formulas from the raw data sheets (cached values are often stale).

    Formula chain:
      K (USDx Balances) → SUMIF across 5 summary sheets col I
        Workday      → SUMIFS on 'Find Bank Statements' (E=date, M=ref, O=bal)
        Lukka        → SUMIFS on 'Daily Snapshot Balance Reconcil' (T=ref, G=bal)
        CCC/CCD      → SUMIFS on 'CCC and CCD Summary' (C=ref, H=bal)
        Bullish      → 'Spot Balance Summary' → EoD Balances (L=H+I+J)
        Balances     → direct values

    Returns JSON:  {"USDx Balances": {"8": 546276258.54, ...}, ...}
    """
    import openpyxl as _openpyxl
    from datetime import datetime as _dt, timedelta as _td
    import re as _re

    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400

    fname = secure_filename(f.filename)
    temp_path = os.path.join(UPLOAD_DIR, f"temp_k_{uuid.uuid4().hex[:8]}_{fname}")
    f.save(temp_path)

    try:
        wb_data = _openpyxl.load_workbook(temp_path, data_only=True)
        wb_form = _openpyxl.load_workbook(temp_path, data_only=False)

        # ----------------------------------------------------------
        # Helper: previous business day (skip weekends)
        # ----------------------------------------------------------
        def _prev_bday(d):
            d = d - _td(days=1)
            while d.weekday() >= 5:
                d = d - _td(days=1)
            return d

        def _to_date(v):
            if isinstance(v, _dt):
                return v.date() if hasattr(v, "date") else v
            return None

        def _to_float(v):
            if isinstance(v, (int, float)):
                return float(v)
            if isinstance(v, str):
                try:
                    return float(v.strip())
                except (ValueError, TypeError):
                    pass
            return None

        # ----------------------------------------------------------
        # Cover Page date
        # ----------------------------------------------------------
        cover_date_raw = None
        if "Cover Page" in wb_data.sheetnames:
            cover_date_raw = wb_data["Cover Page"].cell(6, 2).value
        cover_dt = _to_date(cover_date_raw)  # date object
        prev_bd = _prev_bday(cover_dt) if cover_dt else None

        # ----------------------------------------------------------
        # 1) Workday: read 'Find Bank Statements' raw data
        #    Formula: SUMIFS(O, E=date, M=ref) with MAXIFS date logic
        # ----------------------------------------------------------
        workday_balances = {}  # {ref_float: balance}
        if ("Find Bank Statements" in wb_data.sheetnames
                and "Workday Accts Summary" in wb_data.sheetnames
                and cover_dt):
            ws_fbs = wb_data["Find Bank Statements"]
            # Build {(date, ref_str): sum_O} and {ref_str: max_date}
            fbs_lookup = {}
            fbs_max_date = {}
            for r in range(8, ws_fbs.max_row + 1):
                e_val = ws_fbs.cell(r, 5).value   # E = Statement Date
                m_val = ws_fbs.cell(r, 13).value   # M = Ref number
                o_val = ws_fbs.cell(r, 15).value   # O = Balance
                if m_val is None:
                    continue
                dt = _to_date(e_val)
                if dt is None:
                    continue
                ref_str = str(int(float(m_val))) if isinstance(m_val, (int, float)) else str(m_val).strip()
                bal = float(o_val) if isinstance(o_val, (int, float)) else 0.0
                key = (dt, ref_str)
                fbs_lookup[key] = fbs_lookup.get(key, 0.0) + bal
                if ref_str not in fbs_max_date or dt > fbs_max_date[ref_str]:
                    fbs_max_date[ref_str] = dt

            ws_wd = wb_data["Workday Accts Summary"]
            for r in range(9, ws_wd.max_row + 1):
                b = ws_wd.cell(r, 2).value
                rf = _to_float(b)
                if rf is None:
                    continue
                ref_str = str(int(rf))
                max_dt = fbs_max_date.get(ref_str)
                if max_dt == cover_dt:
                    bal = fbs_lookup.get((cover_dt, ref_str), 0.0)
                else:
                    bal = fbs_lookup.get((prev_bd, ref_str), 0.0)
                workday_balances[rf] = bal

        # ----------------------------------------------------------
        # 2) Lukka: read 'Daily Snapshot Balance Reconcil'
        #    Formula: SUMIFS(G, T=ref)
        # ----------------------------------------------------------
        lukka_balances = {}  # {ref_float: balance}
        if "Daily Snapshot Balance Reconcil" in wb_data.sheetnames:
            ws_ds = wb_data["Daily Snapshot Balance Reconcil"]
            for r in range(9, ws_ds.max_row + 1):
                t_val = ws_ds.cell(r, 20).value  # T = ref
                g_val = ws_ds.cell(r, 7).value   # G = balance
                if t_val is None or g_val is None:
                    continue
                rf = _to_float(t_val)
                bal = _to_float(g_val)
                if rf is None or bal is None:
                    continue
                lukka_balances[rf] = lukka_balances.get(rf, 0.0) + bal

        # ----------------------------------------------------------
        # 3) CCC/CCD: read 'CCC and CCD Summary'
        #    Formula: SUMIFS(H, C=ref)
        # ----------------------------------------------------------
        ccc_balances = {}  # {ref_float: balance}
        if "CCC and CCD Summary" in wb_data.sheetnames:
            ws_ccc = wb_data["CCC and CCD Summary"]
            for r in range(1, ws_ccc.max_row + 1):
                c_val = ws_ccc.cell(r, 3).value  # C = ref
                h_val = ws_ccc.cell(r, 8).value  # H = balance
                if c_val is None or h_val is None:
                    continue
                rf = _to_float(c_val)
                bal = _to_float(h_val)
                if rf is None or bal is None:
                    continue
                ccc_balances[rf] = ccc_balances.get(rf, 0.0) + bal

        # ----------------------------------------------------------
        # 4) Bullish Exchange: EoD Balances → Spot Balance Summary
        #    EoD col L = SUM(H:J) — cached L values are stale, compute
        #    Spot Balance: SUMIFS(computed_L, C=symbol, B=cover_date+2)
        #    Bullish row 9 (ref 59) = Spot Balance Summary C21
        #    Bullish row 10 (ref 60) = hardcoded value
        # ----------------------------------------------------------
        bullish_balances = {}  # {ref_float: balance}
        if ("EoD Balances on Exchange" in wb_data.sheetnames
                and "Spot Balance Summary" in wb_data.sheetnames
                and "Bullish Exchange Accts Summary" in wb_data.sheetnames
                and cover_dt):
            ws_eod = wb_data["EoD Balances on Exchange"]
            target_eod_date = _dt(cover_dt.year, cover_dt.month, cover_dt.day) + _td(days=2)

            # Compute L = H+I+J per row, group by (date, symbol)
            eod_by_sym = {}  # {(date, symbol): sum_of_computed_L}
            for r in range(9, ws_eod.max_row + 1):
                d = ws_eod.cell(r, 2).value
                c = ws_eod.cell(r, 3).value
                h = _to_float(ws_eod.cell(r, 8).value) or 0.0
                i = _to_float(ws_eod.cell(r, 9).value) or 0.0
                j = _to_float(ws_eod.cell(r, 10).value) or 0.0
                if d is None or c is None:
                    continue
                computed_l = h + i + j
                key = (d, c)
                eod_by_sym[key] = eod_by_sym.get(key, 0.0) + computed_l

            # Read Spot Balance Summary B10:B20 for symbol criteria
            ws_sbs = wb_data["Spot Balance Summary"]
            spot_total = 0.0
            for r in range(10, 21):
                sym = ws_sbs.cell(r, 2).value
                if sym:
                    spot_total += eod_by_sym.get((target_eod_date, sym), 0.0)

            # Read Bullish Exchange Accts Summary
            ws_be = wb_data["Bullish Exchange Accts Summary"]
            ws_be_f = wb_form["Bullish Exchange Accts Summary"]
            for r in range(9, ws_be.max_row + 1):
                b = ws_be.cell(r, 2).value
                rf = _to_float(b)
                if rf is None:
                    continue
                i_f = ws_be_f.cell(r, 9).value
                i_d = ws_be.cell(r, 9).value
                if isinstance(i_f, str) and "Spot Balance Summary" in i_f:
                    # This row references Spot Balance Summary C21
                    bullish_balances[rf] = spot_total
                elif isinstance(i_d, (int, float)):
                    bullish_balances[rf] = float(i_d)

        # ----------------------------------------------------------
        # 5) Balances to Confirm Weekly: direct values
        # ----------------------------------------------------------
        btc_balances = {}  # {ref_float: balance}
        if "Balances to Confirm Weekly" in wb_data.sheetnames:
            ws_btc = wb_data["Balances to Confirm Weekly"]
            for r in range(1, ws_btc.max_row + 1):
                b = ws_btc.cell(r, 2).value
                i = ws_btc.cell(r, 9).value
                rf = _to_float(b)
                bal = _to_float(i)
                if rf is not None and bal is not None:
                    btc_balances[rf] = btc_balances.get(rf, 0.0) + bal

        # ----------------------------------------------------------
        # Merge all summary balances into unified lookup
        # ----------------------------------------------------------
        sumif_lookup = {}
        for d in [workday_balances, lukka_balances, ccc_balances,
                   bullish_balances, btc_balances]:
            for ref, bal in d.items():
                sumif_lookup[ref] = sumif_lookup.get(ref, 0.0) + bal

        # ----------------------------------------------------------
        # Process each target sheet (USDx Balances, etc.)
        # ----------------------------------------------------------
        result = {}
        for sheet_name in ["USDx Balances", "Ref Acct Balances Full Summary"]:
            if sheet_name not in wb_data.sheetnames:
                continue

            ws_d = wb_data[sheet_name]
            ws_f = wb_form[sheet_name]
            k_values = {}

            # First pass: compute data-row K values
            for row in range(1, min(ws_d.max_row + 1, 2000)):
                k_cached = ws_d.cell(row, 11).value
                k_formula = ws_f.cell(row, 11).value
                j_cached = ws_d.cell(row, 10).value
                b_val = ws_d.cell(row, 2).value
                is_formula = isinstance(k_formula, str) and k_formula.startswith("=")

                if isinstance(k_cached, (int, float)) and not is_formula:
                    k_values[str(row)] = float(k_cached)
                    continue

                if isinstance(k_cached, _dt) and not is_formula:
                    delta = k_cached - _dt(1899, 12, 30)
                    k_values[str(row)] = delta.days + delta.seconds / 86400.0
                    continue

                if not is_formula:
                    continue

                formula = k_formula

                # --- Pattern: ='Cover Page'!$B$6 ---
                if "Cover Page" in formula and "B$6" in formula:
                    if cover_date_raw is not None:
                        if isinstance(cover_date_raw, (int, float)):
                            k_values[str(row)] = float(cover_date_raw)
                        elif isinstance(cover_date_raw, _dt):
                            delta = cover_date_raw - _dt(1899, 12, 30)
                            k_values[str(row)] = delta.days + delta.seconds / 86400.0
                    continue

                # --- Pattern: SUMIF-based (with or without IF wrapper) ---
                if "SUMIF" in formula:
                    ref = _to_float(b_val)
                    if ref is not None:
                        sumif_total = sumif_lookup.get(ref, 0.0)
                        if "=IF(" in formula:
                            if sumif_total == 0:
                                if isinstance(j_cached, (int, float)):
                                    k_values[str(row)] = float(j_cached)
                            else:
                                k_values[str(row)] = sumif_total
                        else:
                            k_values[str(row)] = sumif_total
                        continue

                # --- Pattern: =SUM(K...) → defer to second pass ---
                if "SUM(" in formula and "K" in formula:
                    continue

                # --- Fallback: use cached value if numeric ---
                if isinstance(k_cached, (int, float)):
                    k_values[str(row)] = float(k_cached)
                elif isinstance(k_cached, _dt):
                    delta = k_cached - _dt(1899, 12, 30)
                    k_values[str(row)] = delta.days + delta.seconds / 86400.0

            # Second pass: SUM rows — process bottom-to-top so
            # dependent totals (e.g. K8=SUM(K98,K136,K141)) resolve
            sum_rows = []
            for row in range(1, min(ws_d.max_row + 1, 2000)):
                if str(row) in k_values:
                    continue
                kf = ws_f.cell(row, 11).value
                if isinstance(kf, str) and "=SUM(" in kf:
                    sum_rows.append(row)

            for row in sorted(sum_rows, reverse=True):
                formula = ws_f.cell(row, 11).value
                # Extract inner part: =SUM(K15:K96) → K15:K96
                m_sum = _re.search(r"SUM\(([^)]+)\)", formula)
                if not m_sum:
                    continue
                inner = m_sum.group(1)
                total = 0.0
                for part in inner.split(","):
                    part = part.strip()
                    if ":" in part:
                        m = _re.match(r"K(\d+):K(\d+)", part)
                        if m:
                            r1, r2 = int(m.group(1)), int(m.group(2))
                            for rr in range(r1, r2 + 1):
                                total += k_values.get(str(rr), 0.0)
                    else:
                        m = _re.match(r"K(\d+)", part)
                        if m:
                            total += k_values.get(str(m.group(1)), 0.0)
                k_values[str(row)] = total

            result[sheet_name] = k_values

        wb_data.close()
        wb_form.close()
        return jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


# ---------------------------------------------------------------------------
# Step 1: Bank of Canada FX rate proxy (replaces Vite dev proxy)
# ---------------------------------------------------------------------------

@app.route("/api/boc-fx")
def boc_fx_proxy():
    """Proxy Bank of Canada Valet API requests to avoid CORS issues."""
    import requests as req
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    url = (
        f"https://www.bankofcanada.ca/valet/observations/group/FX_RATES_DAILY/json"
        f"?start_date={start_date}&end_date={end_date}"
    )
    try:
        resp = req.get(url, timeout=15)
        return Response(resp.content, status=resp.status_code,
                        content_type=resp.headers.get("Content-Type", "application/json"))
    except Exception as e:
        return jsonify({"error": str(e)}), 502


# ---------------------------------------------------------------------------
# Serve the Data Loader HTML for Step 5 (bundled locally)
# ---------------------------------------------------------------------------

@app.route("/api/data-loader-html")
def data_loader_html():
    """Serve the 13WCF Data Loader HTML."""
    loader_path = os.path.join(STEP5_ENGINE_DIR, "data_loader.html")
    if os.path.exists(loader_path):
        with open(loader_path, "r") as f:
            return f.read()
    return "<p>Data Loader HTML not found.</p>", 404


# ---------------------------------------------------------------------------
# Reset workflow
# ---------------------------------------------------------------------------

@app.route("/api/reset", methods=["POST"])
def reset_workflow():
    """Reset all workflow state."""
    workflow["files"].clear()
    workflow["step_outputs"].clear()
    workflow["step_logs"].clear()
    workflow["step_stages"].clear()
    workflow["step_errors"].clear()
    workflow["step_stats"].clear()
    for k in workflow["step_status"]:
        workflow["step_status"][k] = "pending"
    # Clean upload/output dirs
    for d in [UPLOAD_DIR, OUTPUT_DIR]:
        for f in os.listdir(d):
            fp = os.path.join(d, f)
            if os.path.isfile(fp):
                os.remove(fp)
            elif os.path.isdir(fp):
                shutil.rmtree(fp)
    return jsonify({"status": "reset"})


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("=" * 60)
    print("  13WCF Unified Workflow")
    print("  http://localhost:8888")
    print("  Press Ctrl+C to quit")
    print("=" * 60)
    app.run(host="127.0.0.1", port=8888, debug=False, threaded=True)
